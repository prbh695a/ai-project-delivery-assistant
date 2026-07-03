from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def export_project_summaries_to_excel(summaries: list[dict[str, Any]], output_path: Path) -> None:
    """Create Excel output with a real table so Power Automate can read it."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Project Sync Summary"

    headers = [
        "Project ID",
        "Active Ticket Count",
        "Closed Ticket Count",
        "Blocked Ticket Count",
        "Latest Ticket Update",
        "Linked OpenProject Tickets",
        "Last Synced",
    ]
    sheet.append(headers)

    for summary in summaries:
        sheet.append(
            [
                summary["project_id"],
                summary["active_ticket_count"],
                summary["closed_ticket_count"],
                summary["blocked_ticket_count"],
                summary["latest_ticket_update"],
                summary["linked_tickets"],
                summary["last_synced"],
            ]
        )

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 80)

    last_row = max(sheet.max_row, 2)
    if sheet.max_row == 1:
        sheet.append(["", 0, 0, 0, "", "", ""])

    last_column = sheet.max_column
    table_range = f"A1:{get_column_letter(last_column)}{last_row}"
    table = Table(displayName="ProjectSyncTable", ref=table_range)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
